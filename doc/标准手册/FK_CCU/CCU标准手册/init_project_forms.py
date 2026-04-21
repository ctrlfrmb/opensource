from pathlib import Path
import argparse
import datetime as dt
import traceback
import shutil
import openpyxl


DOC_DIR = Path(__file__).resolve().parent
LOG_FILE = DOC_DIR / "init_project_forms.log"


def normalize(s: str) -> str:
    return str(s).strip().replace("：", "").replace(":", "")


def main():
    parser = argparse.ArgumentParser(description="一键初始化 FRM-00~06 项目信息")
    parser.add_argument("--project-id", required=True, help="项目号，例如 2026-001")
    parser.add_argument("--project-name", required=True, help="项目名称")
    parser.add_argument("--product-id", required=True, help="产品编号")
    parser.add_argument("--rig-id", required=True, help="台架编号")
    parser.add_argument("--manager", default="", help="项目经理")
    parser.add_argument("--quality-owner", default="", help="质量负责人")
    parser.add_argument("--customer", default="", help="客户单位（用于FRM-06）")
    parser.add_argument("--env", default="", help="测试环境（用于FRM-03）")
    parser.add_argument("--dut-version", default="", help="DUT版本（用于FRM-03）")
    args = parser.parse_args()

    today = dt.date.today().isoformat()
    logs = []

    def log(msg: str):
        print(msg)
        logs.append(msg)

    key_map = {
        "项目号": args.project_id,
        "项目名称": args.project_name,
        "产品编号": args.product_id,
        "台架编号": args.rig_id,
        "项目经理": args.manager,
        "质量负责人": args.quality_owner,
        "编制日期": today,
        "安装日期": today,
        "测试日期": today,
        "提交日期": today,
        "评审日期": today,
        "验收日期": today,
        "测试环境": args.env,
        "DUT版本": args.dut_version,
        "客户单位": args.customer,
    }

    log(f"[INFO] 开始初始化，目录: {DOC_DIR}")
    # 兼容不同命名风格：FRM-01_xxx.xlsx 或 FRM-01xxx.xlsx
    form_files = sorted(DOC_DIR.glob("FRM-0[0-6]*.xlsx"))
    if not form_files:
        log("[ERROR] 未找到 FRM-00~FRM-06 表单文件，请确认文件在 doc 目录下。")
        DOC_DIR.mkdir(parents=True, exist_ok=True)
        LOG_FILE.write_text("\n".join(logs), encoding="utf-8")
        raise SystemExit(1)

    backup_dir = DOC_DIR / f"backup_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}"
    backup_dir.mkdir(parents=True, exist_ok=True)
    copied = 0
    for f in form_files:
        try:
            shutil.copy2(f, backup_dir / f.name)
            copied += 1
        except Exception as e:
            log(f"[ERROR] 备份失败: {f.name} -> {e}")
            LOG_FILE.write_text("\n".join(logs), encoding="utf-8")
            raise SystemExit(1)
    log(f"[INFO] 备份完成: {backup_dir} (共 {copied} 个文件)")

    updated = []
    failed = []
    for path in form_files:
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            hit_count = 0
            for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=16):
                for c in row:
                    if c.value is None:
                        continue
                    k = normalize(c.value)
                    if k in key_map and c.column < ws.max_column:
                        ws.cell(c.row, c.column + 1, key_map[k])
                        hit_count += 1
            wb.save(path)
            updated.append((path.name, hit_count))
            log(f"[OK] {path.name} 已更新，命中字段 {hit_count} 处")
        except Exception as e:
            failed.append(path.name)
            log(f"[ERROR] {path.name} 更新失败: {e}")
            log(traceback.format_exc())

    log("")
    log("已初始化表单：")
    for f, n in updated:
        log(f"- {f} (字段命中: {n})")

    if failed:
        log("")
        log("失败文件：")
        for f in failed:
            log(f"- {f}")
        log("[ERROR] 存在失败文件，请查看日志并重试。")
        log(f"[INFO] 如需回滚，请从备份目录恢复: {backup_dir}")
    else:
        log("[INFO] 全部表单初始化完成。")
        log(f"[INFO] 备份目录: {backup_dir}")

    DOC_DIR.mkdir(parents=True, exist_ok=True)
    LOG_FILE.write_text("\n".join(logs), encoding="utf-8")
    log(f"[INFO] 日志已写入: {LOG_FILE}")


if __name__ == "__main__":
    main()
